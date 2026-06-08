# cloud_utils.py
import os
import tempfile
from io import BytesIO
from pathlib import Path

import requests
from logger_config import get_logger
from matplotlib.figure import Figure
from requests.exceptions import RequestException, Timeout
from settings import settings
from supabase import create_client
import re
import unicodedata


def secure_filename(filename: str) -> str:
    """Sanitise a filename — drop-in replacement for werkzeug.utils.secure_filename."""
    # Normalise unicode, strip to ASCII, lowercase
    filename = (
        unicodedata.normalize("NFKD", filename)
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    # Replace path separators with underscores
    for sep in (os.sep, os.altsep):
        if sep:
            filename = filename.replace(sep, "_")
    # Keep only alphanumeric, dots, hyphens, underscores
    filename = re.sub(r"[^\w\s\-.]", "", filename).strip()
    # Collapse whitespace
    filename = re.sub(r"\s+", "_", filename)
    return filename or "unnamed"


logger = get_logger(__name__)

# === Environment Setup ===

BASE_DIR = Path(__file__).resolve().parent.parent
IS_PRODUCTION = settings.render == "true"  # Render sets this automatically

# Local directories
IMG_DIR = BASE_DIR / "Outputs" / "Images"
PDF_DIR = BASE_DIR / "Outputs" / "PDFs"
NOTES_DIR = BASE_DIR / "Outputs" / "NOTES"

# Create folders locally if not in production
if not IS_PRODUCTION:
    IMG_DIR.mkdir(parents=True, exist_ok=True)
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    NOTES_DIR.mkdir(parents=True, exist_ok=True)

# === Supabase Setup ===
SUPABASE_URL = settings.supabase_url
SUPABASE_KEY = settings.supabase_key
SUPABASE_BUCKET = "uploads"

supabase = None
if SUPABASE_URL and SUPABASE_KEY:
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    except Exception as e:
        logger.error(f"Failed to initialize Supabase client: {e}")


def sanitize_folder(folder: str) -> str:
    """Sanitize folder name to prevent directory traversal."""
    if not folder:
        return ""
    # Only allow alphanumeric, underscores, and hyphens (and slashes for nested if trusted)
    # For now, let's keep it simple: just basename to be safe, or a whitelist.
    # If the app needs nested folders, we'd need a more complex sanitizer.
    return "/".join(secure_filename(f) for f in folder.split("/") if f)


# === Helper: Upload to Supabase ===
def upload_to_supabase(
    file_bytes: bytes,
    file_name: str,
    folder: str = "",
    content_type: str = "application/pdf",
    session_id: str | None = None,
) -> str:
    """
    Uploads a file to Supabase Storage via direct HTTP POST to set proper content type.
    Returns public URL.
    """
    from database import demo_session_var

    if not session_id:
        session_id = demo_session_var.get()

    if session_id:
        token = demo_session_var.set(session_id)
        try:
            demo_dir = f"/tmp/acatrack_demos/{session_id}"
            os.makedirs(demo_dir, exist_ok=True)
            sanitized_file_name = secure_filename(file_name)
            dest_path = f"{demo_dir}/{sanitized_file_name}"
            with open(dest_path, "wb") as f:
                f.write(file_bytes)
            return "/api/excel/template.xlsx"
        finally:
            demo_session_var.reset(token)

    if not (SUPABASE_URL and SUPABASE_KEY):
        raise RuntimeError("Supabase credentials not loaded.")

    sanitized_file_name = secure_filename(file_name)
    sanitized_folder = sanitize_folder(folder)

    file_path = (
        f"{sanitized_folder}/{sanitized_file_name}"
        if sanitized_folder
        else sanitized_file_name
    )

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": content_type,
        "X-Upsert": "true",
    }

    # URL construction: prevent SSRF/injection by only using configured SUPABASE_URL
    upload_url = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{file_path}"

    res = requests.post(
        upload_url,
        headers=headers,
        data=file_bytes,
        timeout=30,  # Explicit timeout
    )

    if res.status_code not in [200, 201]:
        logger.error(f"Supabase upload failed: {res.status_code} {res.text}")
        raise Exception("Cloud upload failed.")

    public_url = (
        f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{file_path}"
    )
    return public_url


# === Smart File Save ===
def save_file(file, filename: str, folder: str = "files") -> str:
    """Handles both user uploads and generated files (PDFs, images, etc)."""
    filename = secure_filename(filename)
    # folder is sanitized inside upload_to_supabase or here
    file_bytes = file.read() if hasattr(file, "read") else file
    if IS_PRODUCTION:
        content_type = (
            "application/pdf" if filename.lower().endswith(".pdf") else "image/png"
        )
        return upload_to_supabase(
            file_bytes, filename, folder, content_type=content_type
        )
    else:
        # Map folder names to proper paths
        if folder == "files":
            local_folder = IMG_DIR
        elif folder.lower() in ["pdfs"]:
            local_folder = PDF_DIR
        elif folder.lower() in ["images", "imgs"]:
            local_folder = IMG_DIR
        else:
            # Prevent local directory traversal
            local_folder = PDF_DIR / sanitize_folder(folder)

        local_folder.mkdir(parents=True, exist_ok=True)
        save_path = local_folder / filename

        # Verify save_path is still inside Outputs dir
        if not str(save_path.resolve()).startswith(str(BASE_DIR.resolve())):
            raise ValueError("Invalid save location")

        with open(save_path, "wb") as f:
            if hasattr(file, "read"):
                # Warning: if file was already read once, this might fail or be empty
                pass  # already handled via file_bytes logic above
            f.write(file_bytes)
        return str(save_path)


# === Save Matplotlib Figures ===
def save_plot(fig: Figure, filename: str, folder: str = "plots") -> str:
    """Saves matplotlib figures locally or to Supabase Storage."""
    filename = secure_filename(filename)
    if IS_PRODUCTION:
        buf = BytesIO()
        fig.savefig(buf, format="png")
        buf.seek(0)
        return upload_to_supabase(
            buf.read(), filename, folder, content_type="image/png"
        )
    else:
        local_folder = IMG_DIR / sanitize_folder(folder)
        local_folder.mkdir(parents=True, exist_ok=True)
        save_path = local_folder / filename
        fig.savefig(save_path)
        return str(save_path)


def download_excel_from_supabase(
    excel_filename: str, folder: str, session_id: str | None = None
) -> str:
    """Downloads an Excel file from Supabase to a local temp path. Returns local path."""
    from database import demo_session_var

    if not session_id:
        session_id = demo_session_var.get()

    def _download_excel_from_supabase_sandbox(
        excel_filename: str, folder: str, session_id: str
    ) -> str:
        demo_dir = f"/tmp/acatrack_demos/{session_id}"
        os.makedirs(demo_dir, exist_ok=True)
        sanitized_file_name = secure_filename(excel_filename)
        dest_path = f"{demo_dir}/{sanitized_file_name}"
        if os.path.exists(dest_path):
            # Verify if it contains the demo student USN
            try:
                import openpyxl

                wb = openpyxl.load_workbook(dest_path, read_only=True)
                has_demo = False
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    # Check first column (typically USN) in first 30 rows
                    for r in range(1, 31):
                        val1 = sheet.cell(row=r, column=1).value
                        val2 = sheet.cell(row=r, column=2).value
                        if (val1 and "1XX23CS" in str(val1)) or (
                            val2 and "1XX23CS" in str(val2)
                        ):
                            has_demo = True
                            break
                    if has_demo:
                        break
                wb.close()
                if has_demo:
                    return dest_path
            except Exception:
                pass

        import shutil

        scratch_dir = os.path.join(BASE_DIR, "scratch")

        # 1. Dynamic Results Excel Generation
        if "result_list" in excel_filename.lower():
            try:
                from services.student_service import _get_sync_session
                from sqlalchemy import text
                import pandas as pd

                SessionMaker, engine = _get_sync_session()
                with SessionMaker() as session:
                    # Query all subjects
                    subj_rows = session.execute(
                        text(
                            "SELECT subject_code, subject_name, credits, semester FROM subjects"
                        )
                    ).fetchall()
                    # Query all students
                    stud_rows = session.execute(
                        text(
                            "SELECT id, usn, name FROM students WHERE batch_year = :by"
                        ),
                        {"by": int(folder) if folder.isdigit() else 2023},
                    ).fetchall()
                    # Query all academic results
                    res_rows = session.execute(
                        text(
                            "SELECT student_id, subject_code, ia_marks, see_marks FROM academic_results"
                        )
                    ).fetchall()

                # Group subjects by semester
                sem_subjects = {}
                for row in subj_rows:
                    sem = row[3]  # semester
                    if sem not in sem_subjects:
                        sem_subjects[sem] = []
                    sem_subjects[sem].append(row)

                # Index results by (student_id, subject_code)
                res_map = {}
                for r_row in res_rows:
                    res_map[(r_row[0], r_row[1])] = (
                        r_row[2],
                        r_row[3],
                    )  # (ia_marks, see_marks)

                with pd.ExcelWriter(dest_path, engine="openpyxl") as writer:
                    # Generate a sheet for each semester
                    for sem_name, subjs in sem_subjects.items():
                        # Sort subjects by code to keep it deterministic
                        subjs = sorted(subjs, key=lambda x: x[0])

                        # Build columns list
                        cols = ["student_usn", "student_name"]
                        for sub in subjs:
                            code = sub[0]
                            cols.extend(
                                [
                                    f"{code}_INTERNALS",
                                    f"{code}_EXTERNALS",
                                    f"{code}_TOTAL",
                                    f"{code}_CREDITS",
                                ]
                            )

                        # Build row data
                        sheet_rows = []
                        for stud in stud_rows:
                            stud_id, usn, name = stud
                            row_dict = {"student_usn": usn, "student_name": name}
                            for sub in subjs:
                                code, _, creds, _ = sub
                                ia, see = res_map.get((stud_id, code), (0, 0))
                                row_dict[f"{code}_INTERNALS"] = ia
                                row_dict[f"{code}_EXTERNALS"] = see
                                row_dict[f"{code}_TOTAL"] = ia + see
                                row_dict[f"{code}_CREDITS"] = creds
                            sheet_rows.append(row_dict)

                        df = pd.DataFrame(sheet_rows, columns=cols)
                        df.to_excel(writer, sheet_name=sem_name, index=False)

                return dest_path
            except Exception as e:
                logger.error(
                    f"Error dynamically generating demo results excel: {e}",
                    exc_info=True,
                )

        # 2. Dynamic Student Enrollment Excel Generation
        elif "student" in excel_filename.lower() and "enroll" in excel_filename.lower():
            try:
                from services.student_service import _get_sync_session
                from sqlalchemy import text
                import pandas as pd

                SessionMaker, engine = _get_sync_session()
                with SessionMaker() as session:
                    # Query all students
                    stud_rows = session.execute(
                        text(
                            "SELECT usn, name, student_email, student_phno FROM students WHERE batch_year = :by"
                        ),
                        {"by": int(folder) if folder.isdigit() else 2023},
                    ).fetchall()

                sheet_rows = []
                for usn, name, email, phone in stud_rows:
                    sheet_rows.append(
                        {
                            "usn": usn,
                            "name": name,
                            "email": email or "",
                            "phone": phone or "",
                        }
                    )

                df = pd.DataFrame(sheet_rows, columns=["usn", "name", "email", "phone"])
                df.to_excel(dest_path, index=False)
                return dest_path
            except Exception as e:
                logger.error(
                    f"Error dynamically generating student enrollment excel: {e}",
                    exc_info=True,
                )

        # 3. Static templates fallbacks
        fallback_file = os.path.join(scratch_dir, "students_enrollment_2023.xlsx")

        if "subject" in excel_filename.lower():
            fallback_file = os.path.join(scratch_dir, "subjects.xlsx")
        elif "staff" in excel_filename.lower():
            fallback_file = os.path.join(scratch_dir, "staff.xlsx")
        elif "mentor" in excel_filename.lower():
            fallback_file = os.path.join(scratch_dir, "mentor_mapping_2023.xlsx")

        if os.path.exists(fallback_file):
            shutil.copy(fallback_file, dest_path)
            return dest_path

        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        return temp_path

    if session_id:
        token = demo_session_var.set(session_id)
        try:
            return _download_excel_from_supabase_sandbox(
                excel_filename, folder, session_id
            )
        finally:
            demo_session_var.reset(token)

    if not (SUPABASE_URL and SUPABASE_KEY):
        raise RuntimeError("Supabase credentials not loaded.")

    file_path = (
        f"{sanitize_folder(folder)}/{secure_filename(excel_filename)}"
        if folder
        else secure_filename(excel_filename)
    )
    url = (
        f"{SUPABASE_URL}/storage/v1/object/authenticated/{SUPABASE_BUCKET}/{file_path}"
    )

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }

    logger.debug(f"Downloading from Supabase: {url}")
    try:
        res = requests.get(url, headers=headers, timeout=30)
        res.raise_for_status()
    except Exception as e:
        logger.warning(f"Failed authenticated download: {str(e)}")
        public_url = (
            f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{file_path}"
        )
        try:
            res = requests.get(public_url, timeout=30)
            res.raise_for_status()
        except Exception as e2:
            logger.error(f"Failed public download: {str(e2)}")
            raise Exception("Failed to download file from cloud storage.")

    fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    with os.fdopen(fd, "wb") as tmp:
        tmp.write(res.content)
    return temp_path


def excel_exists_in_supabase(excel_filename: str, folder: str) -> bool:
    """Checks if Excel file exists in Supabase Storage."""
    from database import demo_session_var

    if demo_session_var.get():
        return True

    if not supabase:
        return False
    # sanitized_folder = sanitize_folder(folder)
    try:
        files = supabase.storage.from_(SUPABASE_BUCKET).list(folder)
        return secure_filename(excel_filename) in [f["name"] for f in files]
    except Exception as e:
        logger.error(f"Error checking file existence: {str(e)}")
        return False


def upload_pdf_to_supabase(local_pdf_path: str, usn: str, folder: str):
    """Uploads a local PDF to Supabase Storage and returns the public URL."""
    with open(local_pdf_path, "rb") as f:
        return upload_to_supabase(
            f.read(),
            f"{secure_filename(usn)}.pdf",
            folder,
            content_type="application/pdf",
        )


def upload_excel_to_supabase(
    local_excel_path: str,
    excel_filename: str,
    folder: str,
    session_id: str | None = None,
):
    """Uploads a local Excel file to Supabase Storage and returns the public URL."""
    with open(local_excel_path, "rb") as f:
        return upload_to_supabase(
            f.read(),
            excel_filename,
            folder,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            session_id=session_id,
        )


def list_supabase_file_tree(folder: str = "") -> dict:
    """
    Recursively lists folders/files (.pdf only) in Supabase Storage under the specified folder.
    """
    if not (SUPABASE_URL and SUPABASE_KEY and supabase):
        raise RuntimeError("Supabase credentials not loaded.")

    tree = {}
    try:
        files = supabase.storage.from_(SUPABASE_BUCKET).list(folder)
    except Exception as e:
        logger.error(f"Supabase list error: {str(e)}")
        raise RuntimeError("Failed to list cloud storage contents.")

    if not isinstance(files, list):
        return tree

    for entry in files:
        if not entry or not isinstance(entry, dict):
            continue
        name = entry.get("name")
        metadata = entry.get("metadata") or {}
        mimetype = metadata.get("mimetype", "")

        if mimetype == "application/x-directory":
            subfolder = f"{folder}/{name}" if folder else name
            sub_tree = list_supabase_file_tree(subfolder)
            if sub_tree:
                tree[name] = sub_tree
        elif name and name.lower().endswith(".pdf"):
            file_path = f"{folder}/{name}" if folder else name
            url = (
                f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{file_path}"
            )
            tree[name] = url
    return tree


def download_image_from_url(url: str) -> str:
    """
    Downloads an image from a URL to a temporary local file.
    """
    # Potential SSRF: validating domain could be added here
    try:
        # Added explicit timeout and removed redundant call
        headers = {"User-Agent": "AcaTrack-Backend/1.0"}
        response = requests.get(url, timeout=10, headers=headers)
        response.raise_for_status()

        # Create a temporary file safely
        fd, path = tempfile.mkstemp(suffix=".png")
        try:
            with os.fdopen(fd, "wb") as tmp:
                tmp.write(response.content)
        except Exception:
            os.remove(path)
            raise

        return path
    except (RequestException, Timeout) as e:
        logger.warning(f"Could not download image from {url}: {str(e)}")

        fallback_logo = os.path.join(
            os.path.dirname(__file__), "..", "Inputs", "Images", "logo.png"
        )
        return str(Path(fallback_logo).resolve())
