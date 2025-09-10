import re
from pathlib import Path
import pdfplumber
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Logging setup ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- PDF Extraction ---
def extract_text_from_pdf(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    logging.debug(f"Extracted text from {pdf_path}:\n{text[:200]}...")  # Log first 200 chars
    return text

# --- Extract student info and subjects ---
def extract_info(content):
    usn_match = re.search(r'University Seat Number\s*:\s*(\w+)', content, re.IGNORECASE)
    name_match = re.search(r'Student Name\s*:\s*(.+)', content, re.IGNORECASE)
    semester_match = re.search(r'Semester\s*:\s*(\d+)', content, re.IGNORECASE)
    
    usn = usn_match.group(1) if usn_match else "N/A"
    name = name_match.group(1).strip() if name_match else "N/A"
    semester = semester_match.group(1) if semester_match else "N/A"
    
    # Match subjects: Code, Name, Internal, External, Total, Result, Date
    subject_pattern = r'(\w+):?\s*(.*?)\s+(\d+)\s+(\d+)\s+(\d+)\s+([PF])\s+(\d{4}-\d{2}-\d{2})'
    subjects = re.findall(subject_pattern, content)
    
    return usn, name, semester, subjects

# --- Process extracted data for Excel ---
def process_pdf_content(content):
    usn, name, semester, subjects = extract_info(content)
    rows = []
    
    for subject in subjects:
        code, subject_name, internal, external, total, result, date = subject
        
        # Ensure subject code is in correct format
        if not re.match(r'[A-Z]{3}\d{3}', code):
            # Look for code in subject name if missing
            parts = subject_name.split()
            for part in parts:
                if re.match(r'[A-Z]{3}\d{3}', part):
                    code = part
                    subject_name = subject_name.replace(part, "").strip()
                    break
        
        # Remove numbers mistakenly in subject name
        subject_name_filtered = ' '.join([p for p in subject_name.split() if not any(c.isdigit() for c in p)])
        
        rows.append([usn, name, semester, code.strip(), subject_name_filtered.strip(),
                     internal, external, total, result, date])
    
    return rows

# --- Create Excel file ---
def create_excel_file(all_data, output_file):
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    headers = ['Subject Code', 'Subject Name', 'Internal Marks', 'External Marks', 'Total Marks', 'Result', 'Date']
    
    for student_rows in all_data:
        if not student_rows:
            continue
        usn = student_rows[0][0]
        name = student_rows[0][1]
        semester = student_rows[0][2]
        
        sheet = wb.create_sheet(title=usn)
        
        # Styles
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Student info row
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        cell = sheet.cell(row=1, column=1, value=f"USN: {usn}, Name: {name}, Semester: {semester}")
        cell.font = Font(bold=True, size=12)
        cell.alignment = center_align
        sheet.row_dimensions[1].height = 25

        # Header row
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Data rows
        for row_idx, row in enumerate(student_rows, start=3):
            for col_idx, value in enumerate(row[3:], start=1):  # Start from Subject Code
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center_align if col_idx != 2 else left_align
                cell.border = border

        # Adjust column widths
        for column in sheet.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in column)
            adjusted_width = (max_len + 2) * 1.2
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    wb.save(output_file)
    logging.info(f"Excel file saved: {output_file}")

# --- Main function ---
def main():
    input_folder = Path(os.path.abspath("VTU_Results"))  # <-- Update this path
    if not input_folder.exists() or not input_folder.is_dir():
        logging.error(f"Directory not found: {input_folder}")
        return
    
    output_file = input_folder / "VTU_Results.xlsx"
    
    pdf_files = list(input_folder.glob("*.pdf"))
    if not pdf_files:
        logging.warning(f"No PDFs found in {input_folder}")
        return
    
    all_data = []
    for pdf_file in pdf_files:
        logging.info(f"Processing {pdf_file.name}")
        content = extract_text_from_pdf(pdf_file)
        rows = process_pdf_content(content)
        all_data.append(rows)
    
    create_excel_file(all_data, output_file)
    logging.info("All PDFs processed successfully.")

if __name__ == "__main__":
    main()
